import sqlite3
import os
import pandas as pd
import yaml
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
DB_PATH = "data/nifty100.db"


def load_data():

    conn = sqlite3.connect(DB_PATH)

    ratios = pd.read_sql(
        "SELECT * FROM financial_ratios",
        conn,
    )

    market = pd.read_sql(
        "SELECT * FROM market_cap",
        conn,
    )

    profit = pd.read_sql(
        "SELECT * FROM profitandloss",
        conn,
    )

    sectors = pd.read_sql(
        "SELECT * FROM sectors",
        conn,
    )
    companies = pd.read_sql(
        "SELECT * FROM companies",
        conn,
    )
    conn.close()

    return ratios, market, profit, sectors, companies

def load_config():

    with open("config/screener_config.yaml", "r") as file:

        config = yaml.safe_load(file)

    return config

def prepare_data():

    ratios, market, profit, sectors, companies = load_data()

    ratios["financial_year"] = ratios["year"].astype(str).str[:4]

    market["financial_year"] = market["year"].astype(str)

    df = pd.merge(
        ratios,
        market,
        on=["company_id", "financial_year"],
        how="left",
    )
    df.rename(
        columns={
            "year_x": "year",
            "year_y": "market_year",
        },
        inplace=True,
    )
    df = pd.merge(
        df,
        profit[["company_id", "year", "sales", "net_profit"]],
        on=["company_id", "year"],
        how="left",
    )

    df = pd.merge(
        df,
        sectors[["company_id", "broad_sector"]],
        on="company_id",
        how="left",
    )
    df = pd.merge(
        df,
        companies[
            [
                "id",
                "company_name",
                "roce_percentage",
            ]
        ].rename(
            columns={
                "id": "company_id",
            }
        ),
        on="company_id",
        how="left",
    )
    df.drop(
        columns=[
            "id_x",
            "id_y",
            "financial_year",
            "market_year",
        ],
        inplace=True,
    )
    return df

def get_latest_data():
    df = prepare_data()

    # Exclude TTM
    df = df[df["year"] != "TTM"].copy()

    # Convert year to sortable datetime
    df["year_dt"] = pd.to_datetime(df["year"], format="%Y-%m", errors="coerce")

    # Keep the latest record for each company
    latest = (
        df.sort_values("year_dt")
        .groupby("company_id", as_index=False)
        .tail(1)
        .drop(columns="year_dt")
        .reset_index(drop=True)
    )

    return latest

def apply_filters(df, filters):

    data = df.copy()
    if filters is None:
        filters = {}
    for metric in filters:

        value = filters[metric]

        if value is None:
            continue

        if metric == "roe_min":
            data = data[data["return_on_equity_pct"] >= value]

        elif metric == "de_max":

            financial = data["broad_sector"] == "Financials"

            non_financial = data["debt_to_equity"] <= value

            data = data[financial | non_financial]

        elif metric == "fcf_min":
            data = data[data["free_cash_flow_cr"] >= value]

        elif metric == "revenue_cagr_min":
            data = data[data["revenue_cagr_5yr"] >= value]

        elif metric == "pat_cagr_min":
            data = data[data["pat_cagr_5yr"] >= value]

        elif metric == "opm_min":
            data = data[data["operating_profit_margin_pct"] >= value]

        elif metric == "pe_max":
            data = data[data["pe_ratio"] <= value]

        elif metric == "pb_max":
            data = data[data["pb_ratio"] <= value]

        elif metric == "dividend_yield_min":
            data = data[data["dividend_yield_pct"] >= value]

        elif metric == "icr_min":

            icr = pd.to_numeric(
                data["interest_coverage"],
                errors="coerce"
            )

            data = data[
                (icr >= value)
                | (
                    data["interest_coverage"]
                    .astype(str)
                    .str.upper()
                    == "DEBT FREE"
                )
            ]

        elif metric == "market_cap_min":
            data = data[data["market_cap_crore"] >= value]

        elif metric == "net_profit_min":
            data = data[data["net_profit"] >= value]

        elif metric == "eps_cagr_min":
            data = data[data["eps_cagr_5yr"] >= value]

        elif metric == "asset_turnover_min":
            data = data[data["asset_turnover"] >= value]

        elif metric == "sales_min":
            data = data[data["sales"] >= value]

        else:
            continue

    return data


def calculate_fcf_cagr(df, period=5):

    data = df.copy()

    data = data[
        (data["year"] != "TTM")
        & (data["year"] != "PARSE_ERROR")
    ].copy()

    data["year_number"] = (
        data["year"]
        .astype(str)
        .str[:4]
        .astype(int)
    )

    result = []

    for company_id, company_data in data.groupby("company_id"):

        company_data = company_data.sort_values("year_number")

        latest_row = company_data.iloc[-1]

        latest_year = latest_row["year_number"]
        latest_fcf = latest_row["free_cash_flow_cr"]

        target_year = latest_year - period

        old_rows = company_data[
            company_data["year_number"] == target_year
        ]

        fcf_cagr = None

        if len(old_rows) > 0:

            old_fcf = old_rows.iloc[-1]["free_cash_flow_cr"]

            if (
                pd.notna(old_fcf)
                and pd.notna(latest_fcf)
                and old_fcf > 0
                and latest_fcf > 0
            ):

                fcf_cagr = (
                    (latest_fcf / old_fcf) ** (1 / period) - 1
                ) * 100

        result.append(
            {
                "company_id": company_id,
                "fcf_cagr_5yr": fcf_cagr,
            }
        )

    return pd.DataFrame(result)

def add_scoring_metrics(df):

    data = df.copy()

    history = prepare_data()

    fcf_cagr = calculate_fcf_cagr(history)

    data = pd.merge(
        data,
        fcf_cagr,
        on="company_id",
        how="left"
    )

    data["cfo_pat_ratio"] = (
        data["cash_from_operations_cr"] / data["net_profit"]
    )

    data.loc[
        data["net_profit"] == 0,
        "cfo_pat_ratio"
    ] = None

    data["fcf_positive_flag"] = 0

    data.loc[
        data["free_cash_flow_cr"] > 0,
        "fcf_positive_flag"
    ] = 1
    
    return data

def normalize_metric(data, column, inverse=False):

    values = data[column].copy()

    p10 = values.quantile(0.10)
    p90 = values.quantile(0.90)

    values = values.clip(
        lower=p10,
        upper=p90
    )

    minimum = values.min()
    maximum = values.max()

    if maximum == minimum:
        score = pd.Series(
            50,
            index=data.index
        )

    else:
        score = (
            (values - minimum)
            / (maximum - minimum)
        ) * 100

    if inverse:
        score = 100 - score

    score = score.fillna(50)
    return score

def add_sector_relative_score(df):

    data = df.copy()

    data["sector_relative_score"] = 50.0

    for sector in data["broad_sector"].dropna().unique():

        sector_mask = data["broad_sector"] == sector

        sector_scores = data.loc[
            sector_mask,
            "composite_quality_score"
        ]

        minimum_score = sector_scores.min()
        maximum_score = sector_scores.max()

        if maximum_score == minimum_score:

            data.loc[
                sector_mask,
                "sector_relative_score"
            ] = 50.0

        else:

            data.loc[
                sector_mask,
                "sector_relative_score"
            ] = (
                (
                    sector_scores - minimum_score
                )
                / (
                    maximum_score - minimum_score
                )
                * 100
            )

    data["sector_relative_score"] = (
        data["sector_relative_score"].round(2)
    )

    return data

def export_screener(presets, df):
    os.makedirs(
        "output",
        exist_ok=True
    )
    writer = pd.ExcelWriter(
        "output/screener_output.xlsx",
        engine="openpyxl"
    )

    for name in presets:

        if name == "turnaround_watch":
            continue

        result = apply_filters(
            df,
            presets[name]
        )
        if len(result) == 0:
            continue
        columns = [
            "company_id",
            "company_name",
            "broad_sector",
            "return_on_equity_pct",
            "roce_percentage",
            "net_profit_margin_pct",
            "debt_to_equity",
            "interest_coverage",
            "free_cash_flow_cr",
            "revenue_cagr_5yr",
            "pat_cagr_5yr",
            "eps_cagr_5yr",
            "market_cap_crore",
            "pe_ratio",
            "pb_ratio",
            "dividend_yield_pct",
            "sales",
            "net_profit",
            "composite_quality_score",
            "sector_relative_score"
        ]
        result = result[columns]
        result = result.sort_values(
            "composite_quality_score",
            ascending=False
        )

        result = result.reset_index(drop=True)
        result.insert(
            0,
            "Rank",
            range(1, len(result) + 1)
        )

        result.to_excel(
            writer,
            sheet_name=name.replace("_", " ").title()[:31],
            index=False
        )
        sheet = writer.sheets[
            name.replace("_", " ").title()[:31]
        ]
        sheet.freeze_panes = "A2"
        green = PatternFill(
            start_color="C6EFCE",
            end_color="C6EFCE",
            fill_type="solid"
        )

        red = PatternFill(
            start_color="FFC7CE",
            end_color="FFC7CE",
            fill_type="solid"
        )

        bold = Font(
            bold=True
        )

        for cell in sheet[1]:
            cell.font = bold

        for column in sheet.columns:

            length = 0

            letter = get_column_letter(column[0].column)

            for cell in column:

                if cell.value is not None:

                    if len(str(cell.value)) > length:
                        length = len(str(cell.value))

            sheet.column_dimensions[letter].width = length + 2

        thresholds = presets[name]

        for row in range(2, sheet.max_row + 1):

            for cell in sheet[row]:

                header = sheet.cell(row=1, column=cell.column).value

                if header == "return_on_equity_pct" and thresholds.get("roe_min") is not None:

                    if cell.value is not None and cell.value >= thresholds["roe_min"]:
                        cell.fill = green
                    else:
                        cell.fill = red

                elif header == "debt_to_equity" and thresholds.get("de_max") is not None:

                    if cell.value is not None and cell.value <= thresholds["de_max"]:
                        cell.fill = green
                    else:
                        cell.fill = red

                elif header == "free_cash_flow_cr" and thresholds.get("fcf_min") is not None:

                    if cell.value is not None and cell.value >= thresholds["fcf_min"]:
                        cell.fill = green
                    else:
                        cell.fill = red

                elif header == "revenue_cagr_5yr" and thresholds.get("revenue_cagr_min") is not None:

                    if cell.value is not None and cell.value >= thresholds["revenue_cagr_min"]:
                        cell.fill = green
                    else:
                        cell.fill = red

                elif header == "pat_cagr_5yr" and thresholds.get("pat_cagr_min") is not None:

                    if cell.value is not None and cell.value >= thresholds["pat_cagr_min"]:
                        cell.fill = green
                    else:
                        cell.fill = red

                elif header == "pe_ratio" and thresholds.get("pe_max") is not None:

                    if cell.value is not None and cell.value <= thresholds["pe_max"]:
                        cell.fill = green
                    else:
                        cell.fill = red

                elif header == "pb_ratio" and thresholds.get("pb_max") is not None:

                    if cell.value is not None and cell.value <= thresholds["pb_max"]:
                        cell.fill = green
                    else:
                        cell.fill = red

                elif header == "dividend_yield_pct" and thresholds.get("dividend_yield_min") is not None:

                    if cell.value is not None and cell.value >= thresholds["dividend_yield_min"]:
                        cell.fill = green
                    else:
                        cell.fill = red


    writer.close()


def calculate_score(df):

    data = df.copy()

    data = add_scoring_metrics(data)

    data["roe_score"] = normalize_metric(
        data,
        "return_on_equity_pct"
    )

    data["roce_score"] = normalize_metric(
        data,
        "roce_percentage"
    )

    data["npm_score"] = normalize_metric(
        data,
        "net_profit_margin_pct"
    )

    data["fcf_cagr_score"] = normalize_metric(
        data,
        "fcf_cagr_5yr"
    )

    data["cfo_pat_score"] = normalize_metric(
        data,
        "cfo_pat_ratio"
    )

    data["revenue_cagr_score"] = normalize_metric(
        data,
        "revenue_cagr_5yr"
    )

    data["pat_cagr_score"] = normalize_metric(
        data,
        "pat_cagr_5yr"
    )

    data["de_score"] = normalize_metric(
        data,
        "debt_to_equity",
        inverse=True
    )

    data["icr_score"] = normalize_metric(
        data,
        "interest_coverage"
    )

    data["composite_quality_score"] = (
        data["roe_score"] * 0.15
        + data["roce_score"] * 0.10
        + data["npm_score"] * 0.10
        + data["fcf_cagr_score"] * 0.15
        + data["cfo_pat_score"] * 0.10
        + data["fcf_positive_flag"] * 5
        + data["revenue_cagr_score"] * 0.10
        + data["pat_cagr_score"] * 0.10
        + data["de_score"] * 0.10
        + data["icr_score"] * 0.05
    )

    data["composite_quality_score"] = (
        data["composite_quality_score"].round(2)
    )

    data = data.sort_values(
        "composite_quality_score",
        ascending=False
    )

    data = data.reset_index(drop=True)

    return data

