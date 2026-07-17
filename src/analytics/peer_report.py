import os
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

DB_PATH = "data/nifty100.db"

def load_data():

    conn = sqlite3.connect(DB_PATH)

    peer = pd.read_sql(
        "SELECT * FROM peer_percentiles",
        conn,
    )

    ratios = pd.read_sql(
        "SELECT * FROM financial_ratios",
        conn,
    )

    companies = pd.read_sql(
        """
        SELECT
            id,
            company_name
        FROM companies
        """,
        conn,
    )

    sectors = pd.read_sql(
        "SELECT * FROM sectors",
        conn,
    )

    conn.close()

    return peer, ratios, companies, sectors

def prepare_data():

    peer, ratios, companies, sectors = load_data()

    ratios = ratios.copy()

    ratios = ratios[
        ratios["year"] != "TTM"
    ]

    ratios["year"] = pd.to_datetime(
        ratios["year"],
        format="%Y-%m",
    )

    ratios = ratios.sort_values(
        [
            "company_id",
            "year",
        ]
    )

    latest = ratios.drop_duplicates(
        subset="company_id",
        keep="last",
    )

    data = pd.merge(
        latest,
        peer,
        on="company_id",
        how="left",
    )

    data = pd.merge(
        data,
        companies,
        left_on="company_id",
        right_on="id",
        how="left",
    )

    data = pd.merge(
        data,
        sectors[
            [
                "company_id",
                "broad_sector",
                "sub_sector",
            ]
        ],
        on="company_id",
        how="left",
    )

    data.drop(
        columns=[
            "id_x",
            "id_y",
            "broad_sector_y",
            "sub_sector_y",
        ],
        inplace=True,
    )

    data.rename(
        columns={
            "broad_sector_x": "broad_sector",
            "sub_sector_x": "sub_sector",
        },
        inplace=True,
    )
    data["year"] = data["year"].dt.strftime("%Y-%m")

    return data

def export_peer_report(df):

    os.makedirs(
        "output",
        exist_ok=True,
    )

    output = "output/peer_comparison.xlsx"

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:

        sectors = sorted(
            df["broad_sector"].dropna().unique()
        )

        for sector in sectors:

            peer = df[
                df["broad_sector"] == sector
            ].copy()
            print(sector, len(peer)) 
            peer.to_excel(
                writer,
                sheet_name=sector[:31],
                index=False,
            )

    print("peer_comparison.xlsx created.")

def format_workbook():

    workbook = load_workbook(
        "output/peer_comparison.xlsx"
    )

    green = PatternFill(
        fill_type="solid",
        start_color="C6EFCE",
    )

    yellow = PatternFill(
        fill_type="solid",
        start_color="FFEB9C",
    )

    red = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",
    )

    for sheet in workbook.sheetnames:

        ws = workbook[sheet]

        headers = [
            cell.value
            for cell in ws[1]
        ]

        for col in range(
            1,
            ws.max_column + 1,
        ):

            header = headers[col - 1]

            if (
                header is not None
                and "percentile" in header.lower()
            ):

                for row in range(
                    2,
                    ws.max_row + 1,
                ):

                    cell = ws.cell(
                        row=row,
                        column=col,
                    )

                    if cell.value is None:
                        continue

                    if cell.value >= 75:

                        cell.fill = green

                    elif cell.value <= 25:

                        cell.fill = red

                    else:

                        cell.fill = yellow

    workbook.save(
        "output/peer_comparison.xlsx"
    )

    print("Workbook formatted successfully.")
    
if __name__ == "__main__":

    df = prepare_data()
    print(df["broad_sector"].value_counts())
    export_peer_report(df)

    format_workbook()