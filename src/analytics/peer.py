import sqlite3
import os
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import pandas as pd
DB_PATH = "data/nifty100.db"

def load_data():

    conn = sqlite3.connect(DB_PATH)

    ratios = pd.read_sql(
        "SELECT * FROM financial_ratios",
        conn,
    )

    sectors = pd.read_sql(
        "SELECT * FROM sectors",
        conn,
    )

    companies = pd.read_sql(
        """
        SELECT
            id,
            company_name,
            roce_percentage,
            roe_percentage
        FROM companies
        """,
        conn,
    )

    conn.close()

    return ratios, sectors, companies

def prepare_data():

    ratios, sectors, companies = load_data()

    ratios = ratios.copy()

    ratios = ratios[
        ratios["year"] != "TTM"
    ]

    ratios["year"] = pd.to_datetime(
        ratios["year"],
        format="%Y-%m",
    )

    ratios = ratios.sort_values(
        [
            "company_id",
            "year",
        ]
    )

    latest = ratios.drop_duplicates(
        subset="company_id",
        keep="last",
    )

    data = pd.merge(
        latest,
        sectors[
            [
                "company_id",
                "broad_sector",
                "sub_sector",
            ]
        ],
        on="company_id",
        how="left",
    )

    data = pd.merge(
        data,
        companies,
        left_on="company_id",
        right_on="id",
        how="left",
    )

    data.drop(
        columns=[
            "id_x",
            "id_y",
        ],
        inplace=True,
    )

    return data

def calculate_percentile(data, column, ascending=True):

    score = data[column].rank(
        pct=True,
        ascending=ascending,
        na_option="bottom",
    )

    score = score * 100

    return score

def calculate_peer_scores(df):

    data = df.copy()

    groups = data.groupby(
        "broad_sector"
    )

    for sector, group in groups:

        index = group.index

        data.loc[index, "roe_percentile"] = calculate_percentile(
            group,
            "return_on_equity_pct",
        )

        data.loc[index, "roce_percentile"] = calculate_percentile(
            group,
            "roce_percentage",
        )

        data.loc[index, "npm_percentile"] = calculate_percentile(
            group,
            "net_profit_margin_pct",
        )

        data.loc[index, "revenue_cagr_percentile"] = calculate_percentile(
            group,
            "revenue_cagr_5yr",
        )

        data.loc[index, "pat_cagr_percentile"] = calculate_percentile(
            group,
            "pat_cagr_5yr",
        )

        data.loc[index, "eps_cagr_percentile"] = calculate_percentile(
            group,
            "eps_cagr_5yr",
        )

        data.loc[index, "asset_turnover_percentile"] = calculate_percentile(
            group,
            "asset_turnover",
        )

        data.loc[index, "interest_coverage_percentile"] = calculate_percentile(
            group,
            "interest_coverage",
        )

        data.loc[index, "fcf_percentile"] = calculate_percentile(
            group,
            "free_cash_flow_cr",
        )

        data.loc[index, "de_percentile"] = calculate_percentile(
            group,
            "debt_to_equity",
            ascending=False,
        )
    data["year"] = data["year"].dt.strftime("%Y-%m")
    return data

def save_peer_percentiles(df):

    conn = sqlite3.connect(DB_PATH)

    columns = [
        "company_id",
        "broad_sector",
        "sub_sector",
        "roe_percentile",
        "roce_percentile",
        "npm_percentile",
        "revenue_cagr_percentile",
        "pat_cagr_percentile",
        "eps_cagr_percentile",
        "asset_turnover_percentile",
        "interest_coverage_percentile",
        "fcf_percentile",
        "de_percentile",
    ]

    peer = df[columns].copy()

    peer.to_sql(
        "peer_percentiles",
        conn,
        if_exists="replace",
        index=False,
    )

    conn.close()

def get_peer_comparison(df, company):

    sector = df[
        df["company_id"] == company
    ]["broad_sector"].iloc[0]

    peer = df[
        df["broad_sector"] == sector
    ].copy()

    peer = peer.sort_values(
        "composite_quality_score",
        ascending=False,
    )

    peer = peer.reset_index(drop=True)

    return peer

def export_peer_comparison(df):

    os.makedirs(
        "output",
        exist_ok=True,
    )

    writer = pd.ExcelWriter(
        "output/peer_comparison.xlsx",
        engine="openpyxl",
    )

    percentile_columns = [
        "roe_percentile",
        "roce_percentile",
        "npm_percentile",
        "revenue_cagr_percentile",
        "pat_cagr_percentile",
        "eps_cagr_percentile",
        "asset_turnover_percentile",
        "interest_coverage_percentile",
        "fcf_percentile",
        "de_percentile",
    ]

    green = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    yellow = PatternFill(
        start_color="FFEB9C",
        end_color="FFEB9C",
        fill_type="solid"
    )

    red = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )

    amber = PatternFill(
        start_color="FFD966",
        end_color="FFD966",
        fill_type="solid"
    )

    for sector in sorted(df["broad_sector"].dropna().unique()):

        peer = df[
            df["broad_sector"] == sector
        ].copy()

        peer = peer.sort_values(
            "roe_percentile",
            ascending=False,
        )

        peer = peer.reset_index(drop=True)

        peer.to_excel(
            writer,
            sheet_name=sector[:31],
            index=False,
        )

        sheet = writer.sheets[sector[:31]]

        sheet.freeze_panes = "A2"

        sheet.auto_filter.ref = sheet.dimensions

        for cell in sheet[1]:
            cell.font = Font(
                bold=True
            )

        for column in sheet.columns:

            width = 0

            letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                if cell.value is not None:

                    width = max(
                        width,
                        len(str(cell.value))
                    )

            sheet.column_dimensions[
                letter
            ].width = width + 2

        for row in range(2, sheet.max_row + 1):

            for col in range(1, sheet.max_column + 1):

                header = sheet.cell(
                    row=1,
                    column=col
                ).value

                if header not in percentile_columns:
                    continue

                cell = sheet.cell(
                    row=row,
                    column=col
                )

                if cell.value is None:
                    continue

                if cell.value >= 75:

                    cell.fill = green

                elif cell.value >= 25:

                    cell.fill = yellow

                else:

                    cell.fill = red

        if sheet.max_row >= 2:

            for cell in sheet[2]:
                cell.fill = amber

        median_row = sheet.max_row + 1

        sheet.cell(
            row=median_row,
            column=1
        ).value = "Median"

        for col in range(1, sheet.max_column + 1):

            header = sheet.cell(
                row=1,
                column=col
            ).value

            if header in percentile_columns:

                sheet.cell(
                    row=median_row,
                    column=col
                ).value = round(
                    peer[header].median(),
                    2
                )

        for cell in sheet[median_row]:
            cell.font = Font(
                bold=True
            )

    writer.close()

    print(
        "peer_comparison.xlsx created."
    )
if __name__ == "__main__":

    df = prepare_data()

    df = calculate_peer_scores(df)

    save_peer_percentiles(df)

    export_peer_comparison(df)
    it = df[
        df["broad_sector"] == "Information Technology"
    ]

    print(
        it[
            [
                "company_id",
                "return_on_equity_pct",
                "roe_percentile",
            ]
        ].sort_values(
            "roe_percentile",
            ascending=False,
        )
    )